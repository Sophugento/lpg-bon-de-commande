from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
HEADER_FILL = PatternFill("solid", start_color="D9D9D9")
SECTION_FILL = PatternFill("solid", start_color="F0CAD6")
HEADER_FONT = Font(name="Arial", bold=True, size=10)
SECTION_FONT = Font(name="Arial", bold=True, italic=True, size=10)
DATA_FONT = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_section(cell):
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")

def style_data(cell):
    cell.font = DATA_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=False)

# ─── Sheet 1: Produits ───────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Produits"

headers = [
    "Réf.", "Nom FR", "Nom DE", "Type", "Contenant",
    "Prix CHF (HT)", "Prix vente conseillé CHF", "Catégorie",
    "Sous-catégorie", "Promo éligible"
]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    style_header(cell)

ws1.row_dimensions[1].height = 30

# Raw data: None = empty cell, True/False = boolean, numbers = numeric
# Sections marked with "SECTION:" prefix
data = [
    # CORPS — RÉDUCTION DE VOLUME
    ("SECTION", "CORPS — RÉDUCTION DE VOLUME"),
    (102435100, "GEL LIPORÉDUCTEUR", "BODY SHAPING GEL", "revente", "200ml", 35.9, 71.8, "Corps", "Réduction de volume", True),
    (102435200, "GEL LIPORÉDUCTEUR", "BODY SHAPING GEL", "professionnel", "400ml", 55.5, None, "Corps", "Réduction de volume", False),
    (102432100, "CRÈME LIPORÉDUCTEUR", "BODY SHAPING CREAM", "revente", "200ml", 35.9, 71.8, "Corps", "Réduction de volume", True),
    (102432600, "CRÈME LIPORÉDUCTEUR", "BODY SHAPING CREAM", "professionnel", "400ml", 55.5, None, "Corps", "Réduction de volume", False),
    # CORPS — CELLULITE
    ("SECTION", "CORPS — CELLULITE"),
    (102486300, "GEL-CRÈME ANTI CELLULITE", "ANTI-CELLULITE CREAM-IN-GEL", "revente", "200ml", 35.9, 71.8, "Corps", "Cellulite", True),
    (102486400, "GEL-CRÈME ANTI CELLULITE", "ANTI-CELLULITE CREAM-IN-GEL", "professionnel", "400ml", 55.5, None, "Corps", "Cellulite", False),
    (102488900, "SÉRUM INTENSIF ANTI-CELLULITE", "ANTI-CELLULITE INTENSIVE SERUM", "revente", "100ml", 45.5, 91, "Corps", "Cellulite", True),
    (102489000, "SÉRUM INTENSIF ANTI-CELLULITE", "ANTI-CELLULITE INTENSIVE SERUM", "professionnel", "200ml", 62.5, None, "Corps", "Cellulite", False),
    # CORPS — ANTI-AGING
    ("SECTION", "CORPS — ANTI-AGING"),
    (102441500, "CRÈME MICRO PEELING", "RESURFACING CREAM", "revente", "200ml", 35.9, 71.8, "Corps", "Anti-aging", True),
    (102441600, "CRÈME MICRO PEELING", "RESURFACING CREAM", "professionnel", "400ml", 55.5, None, "Corps", "Anti-aging", False),
    (102487100, "CRÈME FERMETÉ GALBANTE", "FIRMING SHAPING CREAM", "revente", "200ml", 35.9, 71.8, "Corps", "Anti-aging", True),
    (102487200, "CRÈME FERMETÉ GALBANTE", "FIRMING SHAPING CREAM", "professionnel", "400ml", 55.5, None, "Corps", "Anti-aging", False),
    # CORPS — BIEN-ÊTRE
    ("SECTION", "CORPS — BIEN-ÊTRE"),
    (102488400, "HUILE EXPERTE VERGETURES", "EXPERT STRETCH MARK OIL", "revente", "100ml", 34, 68, "Corps", "Bien-être", True),
    (102503400, "HUILE EXPERTE VERGETURES", "EXPERT STRETCH MARK OIL", "professionnel", "100ml", 30.5, None, "Corps", "Bien-être", False),
    (102689000, "BAUME DE MODELAGE CRYO", "CRYO MASSAGE BALM", "professionnel", "500ml", 86.5, None, "Corps", "Bien-être", False),
    # VISAGE — NETTOYER & PRÉPARER
    ("SECTION", "VISAGE — NETTOYER & PRÉPARER"),
    (102501000, "BAUME DÉMAQUILLANT", "EXPERT CLEANSING BALM", "revente", "100ml", 25, 50, "Visage", "Nettoyer & Préparer", True),
    (102501100, "BAUME DÉMAQUILLANT", "EXPERT CLEANSING BALM", "professionnel", "200ml", 42, None, "Visage", "Nettoyer & Préparer", False),
    (102435400, "EAU MICELLAIRE PRÉPARATRICE", "PREPARING MICELLAR WATER", "revente", "200ml", 16, 32, "Visage", "Nettoyer & Préparer", True),
    (102435500, "EAU MICELLAIRE PRÉPARATRICE", "PREPARING MICELLAR WATER", "professionnel", "400ml", 30.5, None, "Visage", "Nettoyer & Préparer", False),
    (102434300, "CRÈME EXFOLIANTE", "EXFOLIATING CREAM", "revente", "75ml", 19.5, 39, "Visage", "Nettoyer & Préparer", True),
    (102434400, "CRÈME EXFOLIANTE", "EXFOLIATING CREAM", "professionnel", "200ml", 41, None, "Visage", "Nettoyer & Préparer", False),
    (102501500, "ESSENCE ACTIVE RÉHYDRATANTE", "ACTIVE REPLENISHING ESSENCE", "revente", "200ml", 30, 60, "Visage", "Nettoyer & Préparer", True),
    (102501600, "ESSENCE ACTIVE RÉHYDRATANTE", "ACTIVE REPLENISHING ESSENCE", "professionnel", "400ml", 48, None, "Visage", "Nettoyer & Préparer", False),
    (102589600, "POUDRE SOYEUSE MICRO EXFOLIANTE", "MICRO EXFOLIATING SILKY POWDER", "revente", "50gr", 20.5, 41, "Visage", "Nettoyer & Préparer", True),
    (102339400, "KIT PEELING 20%", "KIT PEELING 20%", "professionnel", None, 48, None, "Visage", "Nettoyer & Préparer", False),
    # VISAGE — SÉRUM
    ("SECTION", "VISAGE — SÉRUM"),
    (102444500, "SÉRUM LACTÉ LISSANT REPULPANT", "SMOOTHING FILLER MILK SERUM", "revente", "30ml", 52, 104, "Visage", "Sérum", True),
    (102444600, "SÉRUM LACTÉ LISSANT REPULPANT", "SMOOTHING FILLER MILK SERUM", "professionnel", "100ml", 83, None, "Visage", "Sérum", False),
    (102499000, "SÉRUM LIFT RAFFERMISSANT", "FIRMING LIFT SERUM", "revente", "30ml", 55, 110, "Visage", "Sérum", True),
    (102499100, "SÉRUM LIFT RAFFERMISSANT", "FIRMING LIFT SERUM", "professionnel", "100ml", 109, None, "Visage", "Sérum", False),
    (102502000, "SÉRUM HUILE-EN-EAU RÉHYDRATANT", "OIL-IN-WATER REPLENISHING SERUM", "revente", "30ml", 46, 92, "Visage", "Sérum", True),
    (102502100, "SÉRUM HUILE-EN-EAU RÉHYDRATANT", "OIL-IN-WATER REPLENISHING SERUM", "professionnel", "100ml", 73, None, "Visage", "Sérum", False),
    (101994200, "SÉRUM ANTI-ÂGE RÉGÉNÉRATION CELLULAIRE", "AGE DEFYING RENEWAL SERUM", "revente", "30ml", 70, 140, "Visage", "Sérum", True),
    (102001500, "SÉRUM ANTI-ÂGE RÉGÉNÉRATION CELLULAIRE", "AGE DEFYING RENEWAL SERUM", "professionnel", "30ml", 60, None, "Visage", "Sérum", False),
    # VISAGE — CRÈMES JOUR/NUIT
    ("SECTION", "VISAGE — CRÈMES JOUR/NUIT"),
    (102442600, "CRÈME LISSANTE REPULPANTE", "SMOOTHING FILLER CREAM", "revente", "50ml", 50, 100, "Visage", "Crèmes Jour / Nuit", True),
    (102442700, "CRÈME LISSANTE REPULPANTE", "SMOOTHING FILLER CREAM", "professionnel", "200ml", 80, None, "Visage", "Crèmes Jour / Nuit", False),
    (102443700, "GEL-CRÈME LISSANTE REPULPANTE", "SMOOTHING FILLER GEL-CREAM", "revente", "50ml", 50, 100, "Visage", "Crèmes Jour / Nuit", True),
    (102443800, "GEL-CRÈME LISSANTE REPULPANTE", "SMOOTHING FILLER GEL-CREAM", "professionnel", "200ml", 80, None, "Visage", "Crèmes Jour / Nuit", False),
    (102498000, "CRÈME LIFT RAFFERMISSANTE", "FIRMING LIFT CREAM", "revente", "50ml", 52, 104, "Visage", "Crèmes Jour / Nuit", True),
    (102498100, "CRÈME LIFT RAFFERMISSANTE", "FIRMING LIFT CREAM", "professionnel", "200ml", 82.5, None, "Visage", "Crèmes Jour / Nuit", False),
    (102503000, "GEL-CRÈME DYNAMISANT RÉHYDRATANT", "DYNAMIC REPLENISHING GEL-CREAM", "revente", "50ml", 41, 82, "Visage", "Crèmes Jour / Nuit", True),
    (102503100, "GEL-CRÈME DYNAMISANT RÉHYDRATANT", "DYNAMIC REPLENISHING GEL-CREAM", "professionnel", "200ml", 66, None, "Visage", "Crèmes Jour / Nuit", False),
    (102589100, "CRÈME RICHE DYNAMISANTE RÉHYDRATANTE", "DYNAMIC REPLENISHING RICH CREAM", "revente", "50ml", 41, 82, "Visage", "Crèmes Jour / Nuit", True),
    (102589200, "CRÈME RICHE DYNAMISANTE RÉHYDRATANTE", "DYNAMIC REPLENISHING RICH CREAM", "professionnel", "200ml", 66, None, "Visage", "Crèmes Jour / Nuit", False),
    (102259900, "SOIN ANTI-ÂGE RÉGÉNÉRATION CELLULAIRE", "ANTI-AGING RENEWAL CREAM", "revente", "50ml", 69, 138, "Visage", "Crèmes Jour / Nuit", True),
    (102260300, "SOIN ANTI-ÂGE RÉGÉNÉRATION CELLULAIRE", "ANTI-AGING RENEWAL CREAM", "professionnel", "50ml", 53, None, "Visage", "Crèmes Jour / Nuit", False),
    (102726000, "SOIN NUIT RÉGÉNÉRANT SUBLIMATEUR", "REGENERATING ENHANCING NIGHT CREAM", "revente", "50ml", 45.5, 91, "Visage", "Crèmes Jour / Nuit", True),
    (102823000, "SOIN NUIT RÉGÉNÉRANT SUBLIMATEUR", "REGENERATING ENHANCING NIGHT CREAM", "recharge", "50ml", 40, None, "Visage", "Crèmes Jour / Nuit", False),
    (102726200, "SOIN NUIT GLYCOLIQUE RESURFAÇANT", "RESURFACING GLYCOLIC NIGHT CREAM", "revente", "50ml", 45.5, 91, "Visage", "Crèmes Jour / Nuit", True),
    (102822900, "SOIN NUIT GLYCOLIQUE RESURFAÇANT", "RESURFACING GLYCOLIC NIGHT CREAM", "recharge", "50ml", 40, None, "Visage", "Crèmes Jour / Nuit", False),
    # VISAGE — YEUX & LÈVRES
    ("SECTION", "VISAGE — YEUX & LÈVRES"),
    (102446300, "CRÈME YEUX", "EYE CREAM", "revente", "15ml", 37, 74, "Visage", "Yeux & Lèvres", True),
    (102446400, "CRÈME YEUX", "EYE CREAM", "professionnel", "75ml", 59, None, "Visage", "Yeux & Lèvres", False),
    (102504000, "BAUME YEUX", "EYE BALM", "revente", "15ml", 31, 62, "Visage", "Yeux & Lèvres", True),
    (102504100, "BAUME YEUX", "EYE BALM", "professionnel", "75ml", 49, None, "Visage", "Yeux & Lèvres", False),
    (102500000, "SOIN REGARD", "EYE LIFT TREATMENT", "revente", "15ml", 40, 80, "Visage", "Yeux & Lèvres", True),
    (102500100, "SOIN REGARD", "EYE LIFT TREATMENT", "professionnel", "15ml", 28, None, "Visage", "Yeux & Lèvres", False),
    (102523000, "SOIN CONTOUR YEUX ET LÈVRES", "EYE & LIP CONTOUR CREAM", "revente", "9ml", 37, 74, "Visage", "Yeux & Lèvres", True),
    (102523100, "SOIN CONTOUR YEUX ET LÈVRES", "EYE & LIP CONTOUR CREAM", "professionnel", "9ml", 25, None, "Visage", "Yeux & Lèvres", False),
    (102164301, "MASQUE CONTOUR YEUX POST-TRAITEMENT", "POST-TREATMENT EYE CONTOUR MASK", "professionnel", "x10", 92.6, None, "Visage", "Yeux & Lèvres", False),
    # VISAGE — RÉGÉNÉRER
    ("SECTION", "VISAGE — RÉGÉNÉRER"),
    (102589400, "MASQUES CRYO DYNA RÉHYDRA", "CRYO DYNA REHYDRA MASKS", "revente", "75ml", 26, 52, "Visage", "Régénérer", True),
    (102595400, "MASQUES CRYO DYNA RÉHYDRA", "CRYO DYNA REHYDRA MASKS", "professionnel", "200ml", 52, None, "Visage", "Régénérer", False),
    (102700700, "MASQUE LÈVRES POST TRAITEMENT", "POST-CARE LIP MASK", "professionnel", "x10", 87.5, None, "Visage", "Régénérer", False),
    (102569900, "MASQUE BIOCELL COLLAGÈNE POST TRAITEMENT", "POST-CARE COLLAGEN MASK", "professionnel", "x10", 126, None, "Visage", "Régénérer", False),
    (102588900, "LOT 10 MASQUES V-LIFT POST-TT", "POST-TREATMENT V-LIFT MASKS", "professionnel", "x10", 98, None, "Visage", "Régénérer", False),
    (102700800, "FLUIDE UV+ DÉFENSE CELLULAIRE", "CELLULAR DEFENSE UV+ FLUID", "revente", "30ml", 28.5, 57, "Visage", "Régénérer", True),
    (102700900, "FLUIDE UV+ DÉFENSE CELLULAIRE", "CELLULAR DEFENSE UV+ FLUID", "professionnel", "30ml", 23, None, "Visage", "Régénérer", False),
    # VISAGE — MAINS
    ("SECTION", "VISAGE — MAINS"),
    (102700300, "CRÈME MAINS RÉGÉNÉRANTE", "REGENERATING HAND CREAM", "revente", "75ml", 21, 42, "Visage", "Mains", True),
    (102700400, "CRÈME MAINS RÉGÉNÉRANTE", "REGENERATING HAND CREAM", "professionnel", "75ml", 17.5, None, "Visage", "Mains", False),
    (102760900, "MASQUE MAINS POST TRAITEMENT", "POST CARE HAND MASK", "professionnel", "x10", 104, None, "Visage", "Mains", False),
    # NUTRICOSMETICS
    ("SECTION", "NUTRICOSMETICS"),
    (102765000, "RÉDUCTEUR D'APPÉTIT", "APPETITE CONTROL", "revente", "56 gélules", 32, 64, "Nutricosmetics", "Nutricosmetics", True),
    (102689100, "CONCENTRÉ DRAINANT", "FLUID MOBILISING CONCENTRATE", "revente", "DOY 156,8gr", 25, 50, "Nutricosmetics", "Nutricosmetics", True),
    (102726500, "STOP PEAU D'ORANGE", "ORANGE PEEL STOP", "revente", "DOY 182gr", 32, 64, "Nutricosmetics", "Nutricosmetics", True),
    (102433100, "BOOSTER DE VITALITÉ", "VITALITY BOOSTER", "revente", "56 gélules", 25, 50, "Nutricosmetics", "Nutricosmetics", True),
    (102448100, "ACIDE HYALURONIQUE", "HYALURONIC ACID", "revente", "28 gélules", 25, 50, "Nutricosmetics", "Nutricosmetics", True),
    (102448101, "CONCENTRÉ BRÛLE-GRAISSES", "FAT-BURNING CONCENTRATE", "revente", "56 gélules", 25, 50, "Nutricosmetics", "Nutricosmetics", True),
    (102448300, "COLLAGÈNE", "COLLAGEN", "revente", "DOY 140gr", 28.2, 56.4, "Nutricosmetics", "Nutricosmetics", True),
    (102448200, "OMÉGA", "OMEGA", "revente", "56 gélules", 25, 50, "Nutricosmetics", "Nutricosmetics", True),
    (102725901, "SOS PETITS ÉCARTS — ÉD. LIMITÉE", "SOS FAT & SUGAR CONTROL ED LIM", "revente", "10 Sticks", 13, 26, "Nutricosmetics", "Nutricosmetics", True),
    (102359502, "THÉ BIO MINCEUR EXPRESS", "EXPRESS ORGANIC SLIMMING TEA", "revente", "28 sachets", 19.8, 39.6, "Nutricosmetics", "Nutricosmetics", True),
    (102781700, "TISANE STRESS SOMMEIL", "STRESS SLEEP HERBAL TEA", "revente", "28 sachets", 12.5, 25, "Nutricosmetics", "Nutricosmetics", True),
    (102811700, "THÉ BIO GLACÉ DÉTOX — ÉD. LIMITÉE", "ORGANIC DETOX ICE TEA", "revente", "28 sachets", 16, 32, "Nutricosmetics", "Nutricosmetics", True),
    (102811800, "THÉ BIO BIEN-ÊTRE — ÉD. LIMITÉE", "WELL-BEING TEA", "revente", "28 sachets", 16, 32, "Nutricosmetics", "Nutricosmetics", True),
    (102830400, "THÉ BIO MINCEUR VRAC — ÉD. LIMITÉE", "ORGANIC SLIMMING TEA BULK", "revente", "180g VRAC", 45, 90, "Nutricosmetics", "Nutricosmetics", True),
    # COSMETOTEXTILES
    ("SECTION", "COSMETOTEXTILES"),
    (102765100, "LOT 5 CORSAIRE SCULPTANT DÉTOX ANTI-CELLULITE — ÉD. LIM.", "SET 5 ANTI-CELLULITE SHAPING SHORTS", "revente", "Lot x5", 110, 44, "Cosmetotextiles", "Cosmetotextiles", False),
    (102840200, "LOT 5 PANTY MINCEUR VENTRE PLAT — ÉD. LIM.", "SET 5 SLIMMING FLAT STOMACH PANTY", "revente", "Lot x5", 110, 44, "Cosmetotextiles", "Cosmetotextiles", False),
    # CONSOMMABLES
    ("SECTION", "CONSOMMABLES"),
    (102211902, "ENDERMOWEAR BLANC — T.1 / M", "ENDERMOWEAR WEISS — GR.1 / M", "professionnel", "T.1 - M", 17.5, None, "Consommables", "Consommables", False),
    (102212002, "ENDERMOWEAR BLANC — T.2 / L", "ENDERMOWEAR WEISS — GR.2 / L", "professionnel", "T.2 - L", 17.5, None, "Consommables", "Consommables", False),
    (102212102, "ENDERMOWEAR BLANC — T.3 / XL", "ENDERMOWEAR WEISS — GR.3 / XL", "professionnel", "T.3 - XL", 17.5, None, "Consommables", "Consommables", False),
    (102212602, "ENDERMOWEAR BLANC — T.4 / XXL", "ENDERMOWEAR WEISS — GR.4 / XXL", "professionnel", "T.4 - XXL", 17.5, None, "Consommables", "Consommables", False),
    (102209800, "ENDERMOWEAR HOMME — GRIS", "ENDERMOWEAR MEN — GRAU", "professionnel", None, 17.5, None, "Consommables", "Consommables", False),
    (101727202, "KIT ENDERMOLOGIE — LIFT 10 ET 20", "KIT ENDERMOLOGIE — LIFT 10 ET 20", "professionnel", None, 20.5, None, "Consommables", "Consommables", False),
    (102260600, "BOX 6x FILTRE", "BOX 6x FILTER", "professionnel", "Box x6", 60, None, "Consommables", "Consommables", False),
    (102593300, "SAC LPG GRAND (x10)", "LPG SHOPPING BAGS LARGE (x10)", "professionnel", "x10", 18, None, "Consommables", "Consommables", False),
    (102593200, "SAC LPG PETIT (x10)", "LPG SHOPPING BAGS SMALL (x10)", "professionnel", "x10 — 20x10x17cm", 15, None, "Consommables", "Consommables", False),
    (1001735, "ROULEAU PROTECTION JETABLE", "EINWEGSCHUTZROLLE", "professionnel", None, 7.5, None, "Consommables", "Consommables", False),
    (102428501, "LINGETTES DÉSINFECTANTES", "DESINFEKTIONSTÜCHER", "professionnel", None, 10, None, "Consommables", "Consommables", False),
    # GOODIES
    ("SECTION", "GOODIES"),
    (102691800, "CABAS LPG / TOTE BAG", "LPG TRAGETASCHE / TOTE BAG", "revente", None, 35, 70, "Goodies", "Goodies", False),
    (102783400, "TROUSSE CORPS / TOILETRY BAG", "KÖRPERTASCHE / TOILETRY BAG", "revente", None, 10, 20, "Goodies", "Goodies", False),
    (102815500, "NUTRI-BOX LPG", "NUTRI-BOX LPG", "revente", None, 10, 20, "Goodies", "Goodies", False),
    (102786200, "MUG LPG", "MUG LPG", "revente", None, 10, 20, "Goodies", "Goodies", False),
    (102783600, "PINCEAU SILICONE", "SILIKONPINSEL", "revente", None, 7.5, 15, "Goodies", "Goodies", False),
    (102818000, "PATCH SILICONE", "SILIKONPATCH", "revente", None, 10, 20, "Goodies", "Goodies", False),
    # TEXTILES
    ("SECTION", "TEXTILES"),
    (101890000, "TUNIQUE NOIRE — S/36", "SCHWARZE TUNIKA — S/36", "professionnel", "S/36", 75, None, "Textiles", "Tenues", False),
    (101890100, "TUNIQUE NOIRE — M/38", "SCHWARZE TUNIKA — M/38", "professionnel", "M/38", 75, None, "Textiles", "Tenues", False),
    (101980200, "TUNIQUE NOIRE — L/40", "SCHWARZE TUNIKA — L/40", "professionnel", "L/40", 75, None, "Textiles", "Tenues", False),
    (101980300, "TUNIQUE NOIRE — XL/42", "SCHWARZE TUNIKA — XL/42", "professionnel", "XL/42", 75, None, "Textiles", "Tenues", False),
    (102240600, "TUNIQUE BLANCHE — S/36", "WEISSE TUNIKA — S/36", "professionnel", "S/36", 75, None, "Textiles", "Tenues", False),
    (102240800, "TUNIQUE BLANCHE — M/38", "WEISSE TUNIKA — M/38", "professionnel", "M/38", 75, None, "Textiles", "Tenues", False),
    (102240900, "TUNIQUE BLANCHE — L/40", "WEISSE TUNIKA — L/40", "professionnel", "L/40", 75, None, "Textiles", "Tenues", False),
    (102241100, "TUNIQUE BLANCHE — XL/42", "WEISSE TUNIKA — XL/42", "professionnel", "XL/42", 75, None, "Textiles", "Tenues", False),
    (1016938, "PEIGNOIR 100% COTON — M", "BADEMANTEL 100% BAUMWOLLE — M", "professionnel", "M", 75, None, "Textiles", "Accessoires", False),
    (1016939, "PEIGNOIR 100% COTON — L", "BADEMANTEL 100% BAUMWOLLE — L", "professionnel", "L", 75, None, "Textiles", "Accessoires", False),
    (1016940, "PEIGNOIR 100% COTON — XL", "BADEMANTEL 100% BAUMWOLLE — XL", "professionnel", "XL", 75, None, "Textiles", "Accessoires", False),
    (101967800, "BADGE LPG", "LPG BADGE", "professionnel", None, 5.5, None, "Textiles", "Accessoires", False),
    (1016942, "DRAP 90x195cm", "TUCH 90x195cm", "professionnel", "90x195cm", 65, None, "Textiles", "Accessoires", False),
    (1017763, "TONGS / FLIPS", "FLIP-FLOPS", "professionnel", "Unisize", 2.95, None, "Textiles", "Accessoires", False),
    # MOBILIER
    ("SECTION", "MOBILIER"),
    (102635600, "ENDERMOTABLE PREMIUM", "ENDERMOTABLE PREMIUM", "professionnel", "L:186/l:72/H:68-96cm", 3880, None, "Mobilier", "Mobilier", False),
    (102485600, "TABLE BLANCHE / LIÈGE", "WEISSE LIEGE", "professionnel", "L:197/l:80/H:62-90cm", 1900, None, "Mobilier", "Mobilier", False),
    (101771301, "MOBICUBE", "MOBICUBE", "professionnel", "L:57/l:54/H:53cm", 950, None, "Mobilier", "Mobilier", False),
    (101771801, "MOBICUBE XL", "MOBICUBE XL", "professionnel", "L:59/l:53/H:197cm", 1500, None, "Mobilier", "Mobilier", False),
]

NUM_COLS = len(headers)
current_row = 2

for row_data in data:
    if row_data[0] == "SECTION":
        # Merge across all columns for section header
        ws1.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=NUM_COLS
        )
        cell = ws1.cell(row=current_row, column=1, value=f"📌 {row_data[1]}")
        style_section(cell)
        ws1.row_dimensions[current_row].height = 20
        current_row += 1
    else:
        ref, nom_fr, nom_de, type_, contenant, prix_ht, prix_vc, cat, sous_cat, promo = row_data
        values = [ref, nom_fr, nom_de, type_, contenant, prix_ht, prix_vc, cat, sous_cat, promo]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            style_data(cell)
            # Format price columns as numbers
            if col_idx in (6, 7) and val is not None:
                cell.number_format = '#,##0.00'
            # Format ref as text-like number (no comma grouping)
            if col_idx == 1:
                cell.number_format = '0'
        current_row += 1

# Column widths
col_widths = [14, 42, 42, 14, 20, 16, 22, 16, 22, 15]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze pane below header
ws1.freeze_panes = "A2"

# Auto-filter on header row
ws1.auto_filter.ref = f"A1:{get_column_letter(NUM_COLS)}1"

# ─── Sheet 2: Offres spéciales ────────────────────────────────────────────────
ws2 = wb.create_sheet("Offres spéciales")

headers2 = ["ID", "Nom FR", "Nom DE", "Prix CHF (HT)", "Description (contenu)", "Cadeau inclus"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    style_header(cell)
ws2.row_dimensions[1].height = 25

offres = [
    (
        "regenerating",
        'Offre Visage "REGENERATING"',
        'Angebot Gesicht "REGENERATING"',
        222,
        "4x SOIN NUIT RÉGÉNÉRANT SUBLIMATEUR + 1 RECHARGE REGENERATING ENHANCING NIGHT CREAM",
        "1x SET MIT 4 GESICHTSMASSAGE-TOOLS",
    ),
    (
        "resurfacing",
        'Offre Visage "RESURFACING"',
        'Angebot Gesicht "RESURFACING"',
        222,
        "4x SOIN NUIT GLYCOLIQUE RESURFAÇANT + 1 RECHARGE RESURFACING GLYCOLIC NIGHT CREAM",
        "1x SET MIT 4 GESICHTSMASSAGE-TOOLS",
    ),
]

for r, row_data in enumerate(offres, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        style_data(cell)
        if col_idx == 4:
            cell.number_format = '#,##0.00'

col_widths2 = [16, 35, 35, 14, 75, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ─── Sheet 3: Descriptions & Images ──────────────────────────────────────────
ws3 = wb.create_sheet("Descriptions & Images")

headers3 = [
    "Nom FR (clé)", "Description FR", "Description DE",
    "Bénéfice 1 FR", "Bénéfice 2 FR", "Bénéfice 3 FR",
    "Bénéfice 1 DE", "Bénéfice 2 DE", "Bénéfice 3 DE",
    "URL Image (lpg-group.com/media/wysiwyg/...)"
]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    style_header(cell)
ws3.row_dimensions[1].height = 30

col_widths3 = [42, 50, 50, 30, 30, 30, 30, 30, 30, 55]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ─── Sheet 4: Mode d'emploi ───────────────────────────────────────────────────
ws4 = wb.create_sheet("Mode d'emploi")

title_cell = ws4.cell(row=1, column=1, value="📋 Mode d'emploi — LPG Bon de commande 2026")
title_cell.font = Font(name="Arial", bold=True, size=13)
title_cell.fill = PatternFill("solid", start_color="D9D9D9")

instructions = [
    ("Réf.", "Numéro de référence produit. Doit être unique dans le tableau Produits."),
    ("Type", 'Valeurs acceptées : "revente", "professionnel" ou "recharge".'),
    ("Promo éligible", "TRUE ou FALSE. La promo 6+1 / 10+2 s'applique uniquement aux produits revente marqués TRUE."),
    ("Prix vente conseillé CHF", "Laisser vide si non applicable (produits professionnels, recharges, consommables, etc.)."),
    ("Ajouter une ligne", "Copier une ligne existante et modifier les valeurs. Ne pas laisser la colonne Réf. vide."),
    ("Supprimer un produit", "Effacer toute la ligne (ne pas laisser de lignes partiellement vides)."),
    ("Mise à jour dans l'app", "Les modifications sont prises en compte dans l'application dans les 5 minutes après import."),
    ("Feuille Descriptions & Images", "Remplir cette feuille pour enrichir les fiches produits affichées dans l'app (facultatif)."),
    ("Feuille Offres spéciales", "Contient les offres groupées avec prix et cadeaux inclus. Modifier selon les promotions en cours."),
]

ws4.cell(row=2, column=1, value="")

for i, (field, desc) in enumerate(instructions, 3):
    field_cell = ws4.cell(row=i, column=1, value=field)
    field_cell.font = Font(name="Arial", bold=True, size=10)
    field_cell.fill = PatternFill("solid", start_color="EBF1DE")

    desc_cell = ws4.cell(row=i, column=2, value=desc)
    desc_cell.font = Font(name="Arial", size=10)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws4.row_dimensions[i].height = 22

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 80

# ─── Save ─────────────────────────────────────────────────────────────────────
output_path = "/Users/sophiehugentobler/Documents/Claude - Winbiz/LPG_BonCommande_2026_v2.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
